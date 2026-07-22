"""Validate and stage bounded files received from a browser."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import zipfile
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any, NoReturn, cast

from openpyxl import load_workbook

from src.core.common.security_limits import (
    MAX_BROWSER_UPLOAD_BYTES,
    MAX_BROWSER_UPLOAD_CELL_LENGTH,
    MAX_BROWSER_UPLOAD_CELLS,
    MAX_BROWSER_UPLOAD_NAME_LENGTH,
    MAX_BROWSER_UPLOAD_ROWS,
    MAX_IMPORT_COLUMNS,
)
from src.core.models.browser_upload_models import (
    BrowserUpload,
    BrowserUploadKind,
    BrowserUploadRequest,
)
from src.core.models.portfolio_models import PortfolioData
from src.core.models.portfolio_bundle_models import PortfolioBundleContents
from src.core.services.import_preview_service import ImportPreviewService
from src.core.services.portfolio_migrator import PortfolioMigrator
from src.core.services.portfolio_integrity_service import PortfolioIntegrityService
from src.core.services.portfolio_bundle_service import PortfolioBundleService

_MEDIA_TYPES: dict[BrowserUploadKind, frozenset[str]] = {
    "csv": frozenset(
        {
            "",
            "application/csv",
            "application/octet-stream",
            "application/vnd.ms-excel",
            "text/csv",
            "text/plain",
        }
    ),
    "json": frozenset(
        {"", "application/json", "application/octet-stream", "text/json", "text/plain"}
    ),
    "excel": frozenset(
        {
            "",
            "application/octet-stream",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
    ),
    "portfolio": frozenset(
        {"", "application/json", "application/octet-stream", "text/json", "text/plain"}
    ),
    "bundle": frozenset(
        {
            "",
            "application/octet-stream",
            "application/vnd.ring5.portfolio-bundle+zip",
            "application/zip",
        }
    ),
}


def _invalid_json_constant(value: str) -> NoReturn:
    raise ValueError(f"JSON contains unsupported numeric value {value!r}.")


def _unique_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"JSON object contains duplicate key {key!r}.")
        result[key] = value
    return result


def _parse_json(content: bytes) -> Any:
    """Decode strict UTF-8 JSON with duplicate-key and numeric validation."""
    try:
        text = content.decode("utf-8-sig")
        return json.loads(
            text,
            object_pairs_hook=_unique_object,
            parse_constant=_invalid_json_constant,
        )
    except UnicodeDecodeError as exc:
        raise ValueError("JSON uploads must use UTF-8 text encoding.") from exc
    except RecursionError as exc:
        raise ValueError("JSON upload is nested too deeply.") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON parsing failed at line {exc.lineno}, column {exc.colno}.") from exc


def _cell(value: Any) -> str | int | float | bool | None:
    """Normalize one flat tabular value and reject nested or oversized cells."""
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and not math.isfinite(value):
            raise ValueError("Tabular uploads do not support non-finite numeric values.")
        if len(str(value)) > MAX_BROWSER_UPLOAD_CELL_LENGTH:
            raise ValueError(
                f"Upload cell exceeds the {MAX_BROWSER_UPLOAD_CELL_LENGTH:,}-character limit."
            )
        return value
    if hasattr(value, "isoformat"):
        normalized = str(value.isoformat())
        if len(normalized) <= MAX_BROWSER_UPLOAD_CELL_LENGTH:
            return normalized
    raise ValueError("Tabular uploads support only scalar cell values.")


def _json_records(value: Any) -> tuple[tuple[str, ...], list[list[Any]]]:
    """Convert a JSON record object or array into a bounded rectangular table."""
    if isinstance(value, dict) and set(value) == {"records"}:
        value = value["records"]
    if isinstance(value, dict):
        records: list[Any] = [value]
    elif isinstance(value, list):
        records = value
    else:
        raise ValueError(
            "Dataset JSON must be an object, an array of objects, or a records object."
        )
    if not records:
        raise ValueError("Dataset JSON contains no records.")
    if len(records) > MAX_BROWSER_UPLOAD_ROWS:
        raise ValueError(f"Browser uploads support at most {MAX_BROWSER_UPLOAD_ROWS:,} rows.")
    if any(not isinstance(record, dict) for record in records):
        raise ValueError("Every dataset JSON record must be an object.")

    columns: list[str] = []
    for record in records:
        for key in cast(dict[str, Any], record):
            if key not in columns:
                columns.append(key)
                if len(columns) > MAX_IMPORT_COLUMNS:
                    raise ValueError(
                        f"Browser uploads support at most {MAX_IMPORT_COLUMNS:,} columns."
                    )
    if not columns or any(not column.strip() for column in columns):
        raise ValueError("Dataset JSON must contain non-empty column names.")
    if len(records) * len(columns) > MAX_BROWSER_UPLOAD_CELLS:
        raise ValueError(f"Browser uploads support at most {MAX_BROWSER_UPLOAD_CELLS:,} cells.")
    rows = [
        [_cell(cast(dict[str, Any], record).get(column)) for column in columns]
        for record in records
    ]
    return tuple(columns), rows


def _excel_records(content: bytes) -> tuple[tuple[str, ...], list[list[Any]], str]:
    """Read the first visible worksheet from a validated modern Excel workbook."""
    if not zipfile.is_zipfile(io.BytesIO(content)):
        raise ValueError("Excel upload is not a valid .xlsx archive.")
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise ValueError("Excel upload could not be parsed as a safe .xlsx workbook.") from exc
    try:
        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if not visible:
            raise ValueError("Excel upload has no visible worksheet.")
        sheet = visible[0]
        records = sheet.iter_rows(values_only=True)
        try:
            raw_header = next(records)
        except StopIteration as exc:
            raise ValueError("Excel worksheet is empty.") from exc
        header = tuple("" if value is None else str(value).strip() for value in raw_header)
        while header and not header[-1]:
            header = header[:-1]
        if not header or len(header) > MAX_IMPORT_COLUMNS:
            raise ValueError(
                f"Excel header must contain from 1 through {MAX_IMPORT_COLUMNS} columns."
            )
        if any(not value for value in header) or len(set(header)) != len(header):
            raise ValueError("Excel header names must be non-empty and unique.")
        rows: list[list[Any]] = []
        for raw_row in records:
            if len(rows) >= MAX_BROWSER_UPLOAD_ROWS:
                raise ValueError(
                    f"Browser uploads support at most {MAX_BROWSER_UPLOAD_ROWS:,} rows."
                )
            relevant = raw_row[: len(header)]
            if not any(value is not None for value in relevant):
                continue
            if len(rows) * len(header) >= MAX_BROWSER_UPLOAD_CELLS:
                raise ValueError(
                    f"Browser uploads support at most {MAX_BROWSER_UPLOAD_CELLS:,} cells."
                )
            rows.append([_cell(value) for value in relevant])
        if not rows:
            raise ValueError("Excel worksheet contains no data rows.")
        return header, rows, sheet.title
    finally:
        workbook.close()


def _write_table(path: Path, columns: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    """Write a deterministic normalized CSV and enforce its final byte bound."""
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(columns)
    writer.writerows(rows)
    encoded = buffer.getvalue().encode("utf-8")
    if len(encoded) > MAX_BROWSER_UPLOAD_BYTES:
        raise ValueError("Normalized upload exceeds the browser-upload size limit.")
    path.write_bytes(encoded)


def _validate_csv_table(content: bytes, encoding: str, delimiter: str, header_row: int) -> None:
    """Apply browser-specific row, cell-count, and cell-length limits to CSV content."""
    text = content.decode(encoding)
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter, strict=True)
    rows = 0
    cells = 0
    try:
        for record_index, record in enumerate(reader, start=1):
            if any(len(value) > MAX_BROWSER_UPLOAD_CELL_LENGTH for value in record):
                raise ValueError(
                    f"Upload cell exceeds the {MAX_BROWSER_UPLOAD_CELL_LENGTH:,}-character limit."
                )
            if record_index <= header_row or not record:
                continue
            rows += 1
            cells += len(record)
            if rows > MAX_BROWSER_UPLOAD_ROWS:
                raise ValueError(
                    f"Browser uploads support at most {MAX_BROWSER_UPLOAD_ROWS:,} rows."
                )
            if cells > MAX_BROWSER_UPLOAD_CELLS:
                raise ValueError(
                    f"Browser uploads support at most {MAX_BROWSER_UPLOAD_CELLS:,} cells."
                )
    except csv.Error as exc:
        raise ValueError(f"CSV upload parsing failed near line {reader.line_num}: {exc}.") from exc


def _portfolio(
    value: Any,
    *,
    signing_key: str | bytes | None = None,
    require_signature: bool = False,
) -> dict[str, Any]:
    """Validate and migrate a recognizable RING-5 portfolio object."""
    if not isinstance(value, dict):
        raise ValueError("Portfolio JSON must contain one top-level object.")
    if "plots" not in value or not ({"data_csv", "schema_version"} & set(value)):
        raise ValueError("JSON does not contain a recognizable RING-5 portfolio.")
    if not isinstance(value.get("plots"), list):
        raise ValueError("Portfolio plots must be an array.")
    if "data_csv" in value and not isinstance(value["data_csv"], str):
        raise ValueError("Portfolio data_csv must be text.")
    if "config" in value and not isinstance(value["config"], dict):
        raise ValueError("Portfolio config must be an object.")
    try:
        integrity = PortfolioIntegrityService.verify(value, signing_key=signing_key)
        PortfolioIntegrityService.require_restorable(
            integrity,
            require_signature=require_signature,
        )
        return PortfolioMigrator.migrate(cast(dict[str, Any], value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Portfolio schema validation failed: {exc}") from exc


def _kind(file_name: str, value: Any | None, request: BrowserUploadRequest) -> BrowserUploadKind:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        detected: BrowserUploadKind = "csv"
    elif suffix == ".xlsx":
        detected = "excel"
    elif suffix == ".json":
        is_portfolio = (
            isinstance(value, dict)
            and "plots" in value
            and bool({"data_csv", "schema_version"} & set(value))
        )
        detected = "portfolio" if is_portfolio else "json"
    elif suffix == ".ring5-bundle":
        detected = "bundle"
    else:
        raise ValueError("Upload filename must end in .csv, .json, .xlsx, or .ring5-bundle.")
    if request == "portfolio" and suffix != ".json":
        raise ValueError("Portfolio uploads must be JSON files.")
    if request == "portfolio":
        return "portfolio"
    if request == "bundle" and suffix != ".ring5-bundle":
        raise ValueError("Portable bundle uploads must be .ring5-bundle files.")
    if request == "bundle":
        return "bundle"
    if request == "dataset" and detected == "portfolio":
        return "json"
    if request == "dataset" and detected == "bundle":
        raise ValueError("A portable bundle cannot be interpreted as a dataset.")
    return detected


class BrowserUploadService:
    """Inspect browser bytes without changing workspace data or plots."""

    @staticmethod
    def inspect(
        file_name: str,
        content_type: str,
        content: bytes,
        destination: Path,
        request: BrowserUploadRequest = "auto",
    ) -> BrowserUpload:
        # [impl->req~ring5.ingestion.browser-upload~1]
        """Validate, fingerprint, stage, and parse one bounded upload."""
        if request not in {"auto", "dataset", "portfolio", "bundle"}:
            raise ValueError(f"Unsupported browser upload interpretation: {request!r}.")
        if not isinstance(file_name, str):
            raise ValueError("Upload filename is invalid or unsafe.")
        clean_name = Path(file_name).name
        if (
            not clean_name
            or clean_name != file_name
            or "\x00" in clean_name
            or any(ord(character) < 32 for character in clean_name)
            or len(clean_name) > MAX_BROWSER_UPLOAD_NAME_LENGTH
        ):
            raise ValueError("Upload filename is invalid or unsafe.")
        if not isinstance(content, bytes) or not content:
            raise ValueError("Browser upload is empty.")
        if len(content) > MAX_BROWSER_UPLOAD_BYTES:
            raise ValueError(
                f"Browser upload exceeds the {MAX_BROWSER_UPLOAD_BYTES // (1024 * 1024)} MiB limit."
            )

        json_value = _parse_json(content) if Path(clean_name).suffix.lower() == ".json" else None
        kind = _kind(clean_name, json_value, request)
        normalized_media = content_type.partition(";")[0].strip().lower()
        if normalized_media not in _MEDIA_TYPES[kind]:
            raise ValueError(
                f"Declared media type {normalized_media!r} does not match {kind} upload."
            )

        digest = hashlib.sha256(content).hexdigest()
        destination.mkdir(parents=True, exist_ok=True)
        source = destination / f"{digest}{Path(clean_name).suffix.lower()}"
        source.write_bytes(content)

        if kind == "bundle":
            bundle_info = PortfolioBundleService.inspect(content)
            return BrowserUpload(
                file_name=clean_name,
                content_type=normalized_media or "not provided",
                kind=kind,
                size_bytes=len(content),
                source_sha256=digest,
                source_path=str(source.resolve()),
                bundle_info=bundle_info,
            )

        if kind == "portfolio":
            portfolio = _portfolio(json_value)
            integrity = PortfolioIntegrityService.verify(cast(dict[str, Any], json_value))
            return BrowserUpload(
                file_name=clean_name,
                content_type=normalized_media or "not provided",
                kind=kind,
                size_bytes=len(content),
                source_sha256=digest,
                source_path=str(source.resolve()),
                portfolio_schema_version=int(portfolio["schema_version"]),
                portfolio_plot_count=len(portfolio.get("plots", [])),
                portfolio_has_data=bool(portfolio.get("data_csv")),
                portfolio_integrity_status=integrity.status,
                portfolio_signing_key_id=integrity.key_id,
            )

        sheet_name: str | None = None
        if kind == "csv":
            import_path = source
            preview = ImportPreviewService.preview(str(import_path))
            _validate_csv_table(
                content,
                preview.encoding,
                preview.delimiter,
                preview.options.header_row,
            )
            columns = tuple(column.name for column in preview.columns)
            row_count = preview.total_row_count
        else:
            if kind == "json":
                columns, rows = _json_records(json_value)
            else:
                columns, rows, sheet_name = _excel_records(content)
            import_path = destination / f"{digest}.normalized.csv"
            _write_table(import_path, columns, rows)
            preview = ImportPreviewService.preview(str(import_path))
            row_count = preview.total_row_count
        return BrowserUpload(
            file_name=clean_name,
            content_type=normalized_media or "not provided",
            kind=kind,
            size_bytes=len(content),
            source_sha256=digest,
            source_path=str(source.resolve()),
            import_path=str(import_path.resolve()),
            columns=columns,
            row_count=row_count,
            sheet_name=sheet_name,
        )

    @staticmethod
    def load_portfolio(
        upload: BrowserUpload,
        *,
        signing_key: str | bytes | None = None,
        require_signature: bool = False,
    ) -> PortfolioData:
        # [impl->req~ring5.ingestion.browser-upload~1]
        # [impl->req~ring5.portfolio.signed-manifests~1]
        """Revalidate an unchanged staged portfolio for confirmed restoration."""
        if upload.kind != "portfolio":
            raise ValueError("Only a portfolio upload can restore a workspace.")
        content = Path(upload.source_path).read_bytes()
        if hashlib.sha256(content).hexdigest() != upload.source_sha256:
            raise ValueError("Uploaded portfolio changed after validation; upload it again.")
        return cast(
            PortfolioData,
            _portfolio(
                _parse_json(content),
                signing_key=signing_key,
                require_signature=require_signature,
            ),
        )

    @staticmethod
    def load_portfolio_bundle(
        upload: BrowserUpload,
        *,
        signing_key: str | bytes | None = None,
        require_signature: bool = False,
    ) -> PortfolioBundleContents:
        # [impl->req~ring5.portfolio.portable-bundles~1]
        """Revalidate staged bundle bytes and return all verified contents."""
        if upload.kind != "bundle":
            raise ValueError("Only a portable bundle upload can use bundle restoration.")
        content = Path(upload.source_path).read_bytes()
        if hashlib.sha256(content).hexdigest() != upload.source_sha256:
            raise ValueError("Uploaded portable bundle changed after validation; upload it again.")
        return PortfolioBundleService.read(
            content,
            signing_key=signing_key,
            require_signature=require_signature,
        )
